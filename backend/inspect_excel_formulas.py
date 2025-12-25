
from openpyxl import load_workbook
import json

def inspect_formulas():
    file_path = r"d:\Document\SOURCE\PLANNING-PURCHASE\doc\XNT - WIN GỐI_Update 14.12.2025.xlsx"
    sheets_to_check = ['WEEK', 'ORDER']
    
    analysis = {}
    
    try:
        print(f"Loading Workbook (Formula Mode)...")
        wb = load_workbook(file_path, data_only=False)
        
        for sheet_name in sheets_to_check:
            if sheet_name not in wb.sheetnames:
                continue
                
            print(f"Inspecting formulas in {sheet_name}...")
            ws = wb[sheet_name]
            
            # Extract formulas from a sample row (e.g., Row 10 - adjusted based on previous finding that header is ~Row 5-8?)
            # Let's inspect a range of rows to capture the repeating logic
            row_start = 10 
            row_end = 12
            
            sheet_formulas = {}
            for row in range(row_start, row_end + 1):
                row_data = {}
                for col in range(1, 50): # Inspect first 50 columns
                    cell = ws.cell(row=row, column=col)
                    val = cell.value
                    if isinstance(val, str) and val.startswith('='):
                        row_data[f"Col_{col}"] = val
                if row_data:
                    sheet_formulas[f"Row_{row}"] = row_data
            
            analysis[sheet_name] = sheet_formulas
            
        with open("backend/formulas_output.json", "w", encoding="utf-8") as f:
            json.dump(analysis, f, indent=2, ensure_ascii=False)
        print("Done. Saved to backend/formulas_output.json")
        
    except Exception as e:
        print("Error:", e)

if __name__ == "__main__":
    inspect_formulas()
