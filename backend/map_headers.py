
import pandas as pd
import json
from openpyxl.utils import get_column_letter

def map_headers():
    file_path = r"d:\Document\SOURCE\PLANNING-PURCHASE\doc\XNT - WIN GỐI_Update 14.12.2025.xlsx"
    sheet_name = 'WEEK'
    
    print(f"Mapping Headers for Sheet: {sheet_name}")
    
    try:
        # 1. Get Headers
        # Based on inspection, header is likely around row 5-8. 
        # Using the previous logic to find the header row
        df_detect = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=15)
        header_row_idx = 0
        max_non_null = 0
        for i in range(len(df_detect)):
            non_null = df_detect.iloc[i].count()
            if non_null > max_non_null:
                max_non_null = non_null
                header_row_idx = i
                
        print(f"Detected Header Row: {header_row_idx}")
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_idx, nrows=0)
        
        mapping = {}
        for idx, col_name in enumerate(df.columns):
            col_letter = get_column_letter(idx + 1)
            # Clean name
            clean_name = str(col_name).replace('\n', ' ').strip()
            mapping[col_letter] = clean_name
            
        # 2. Load Formulas (Manual load from previous step output or re-extract)
        # For simplicity, let's look at the generated keys from previous step (Col_8, Col_9...)
        # Col_8 = Column H
        # Col_9 = Column I
        
        print("--- COLUMN MAPPING ---")
        # Print first 50 columns to match our formula inspection
        output_list = []
        for i in range(1, 51):
            letter = get_column_letter(i)
            name = mapping.get(letter, "Unknown")
            output_list.append(f"Col_{i} ({letter}): {name}")
            
        print("\n".join(output_list))
        
        # Save to file
        with open("backend/header_mapping.txt", "w", encoding="utf-8") as f:
            f.write("\n".join(output_list))
            
    except Exception as e:
        print("Error:", e)

if __name__ == "__main__":
    map_headers()
